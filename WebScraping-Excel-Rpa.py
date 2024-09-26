import requests ## Para hacer solicitudes HTTP
from bs4 import BeautifulSoup ## Para analizar el contenido HTML de una página web
import openpyxl ## Para crear y manipular archivos de Excel
from openpyxl.styles import Font 
import smtplib ## Para enviar correos electrónicos mediante SMTP
import os ## Para manejar carpetas del sistema operativo

## Para crear mensajes de correo electrónico
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


# Definir la carpeta que se usara
os.chdir (r"C:\Users\David\Desktop\Proyecto-Python-Final")

# url de donde se sacaran los datos
main_url = "https://listado.mercadolibre.com.co/"
print("\n****************************")
busqueda_producto = input("Ingrese lo que desea buscar: ")
print("\n****************************")
url = main_url+""+busqueda_producto+"#D[A:"+busqueda_producto+",L:undefined]"

#Se hace una peticion http
response = requests.get(url) ## si es == 200 continua


#Creamos un archivo de excel y agregamos una hoja
libro = openpyxl.Workbook()
hoja = libro.active
hoja.title = "Productos-MercadoLibre"


#Creamos estilos a las letras
Negrita = Font(bold=True, size="13") 
Tamaño = Font(italic=True, size="15" )
Rojo_Negrita = Font(bold = True, color = "ff2f03" , size=14)

# Creamos los encabezados de las columnas
hoja.cell(row=1, column=1).value = "Producto"
hoja.cell(row=1, column=2).value = "Precio Producto"

# Aplicamos distintos estilos a los titulos
hoja.cell(row=1, column=1).font = Rojo_Negrita
hoja.cell(row=1, column=2).font = Rojo_Negrita

# Tamaño del ancho de la columna
hoja.column_dimensions['A'].width = 100
hoja.column_dimensions['B'].width = 30



if response.status_code == 200:
    soup = BeautifulSoup(response.text, 'html.parser')

    # Seleccionamos todos los productos
    productos = soup.find_all('div', class_="ui-search-result__wrapper")

    fila = 2 #Se inicia desde la segunda fila porque en la primera estan los titulos

    for producto in productos:
        # Titulo del producto
        titulo = producto.find('h2', class_="poly-box poly-component__title")

        # Precio del producto
        precio = producto.find ('span', class_="andes-money-amount__fraction")
        
        # Guardamos el título y el precio en el archivo de excel
        if titulo:
            hoja.cell(row=fila, column=1).value = titulo.text.strip()
            hoja.cell(row=fila, column=1).font = Negrita
        else:
            hoja.cell(row=fila, column=1).value = "Título no encontrado"

        if precio: 
            hoja.cell(row=fila, column=2).value = precio.text.strip()
            hoja.cell(row=fila, column=2).font = Tamaño
        else:
            hoja.cell(row=fila, column=1).value = "Precio No encontrado"

        fila += 1 #Se hace un contador para que vaya pasando de fila y se vaya almacenando

        #Ajustamos el alto de las filas
        for i in range(2, fila):
            hoja.row_dimensions[i].height = 30

    #Guardamos el archivo de excel
    libro.save("Productos-MercadoLibre.xlsx")
    
    print("La busqueda se realizo y fue almacenada correctamente") 
    print("\n****************************")

## si no hay titulo o precio muestra:
else:
    print("Error al cargar la web, codigo: ", response.status_code)


##  Mensaje del correo electronico
mensaje = MIMEMultipart()
print("\n****************************")
cuerpo_mensaje = input("Ingrese el mensaje que contendra el correo: ")

# datos para envar el correo
remitente = 'productos.mercadolibre2024@gmail.com'
contra = 'bakx bxel esot kmok' ## contraseña de aplicacion generada
ruta_adjunto = 'Productos-MercadoLibre.xlsx'
correos_entrada = input("Ingrese los destinatarios del correo, separados por una coma: ")
asunto = input("Ingrese el asunto del correo: ")
print("\n****************************")

destinatarios = [correo.strip() for correo in correos_entrada.split(",")]

# Establecer los atributos del mensaje
mensaje['From'] = remitente
mensaje['To'] = ", ".join(destinatarios) ## cadena para varios destinatarios
mensaje['Subject'] = asunto

# Agregar el cuerpo del mensaje
mensaje.attach(MIMEText(cuerpo_mensaje, 'plain')) ## texto tipo plano

# Adjuntar archivo
with open(ruta_adjunto, 'rb') as archivo_adjunto: ## abre archivo en lectura binaria rb
    adjunto_MIME = MIMEBase('application', 'octet-stream') ## creacion de instancia para adjuntar archivos
    adjunto_MIME.set_payload(archivo_adjunto.read()) ## se carga el contenido del xlsx
    encoders.encode_base64(adjunto_MIME) 
    adjunto_MIME.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(ruta_adjunto)}") ## Se agrega un encabezado que indica que este contenido es un archivo adjunto
    mensaje.attach(adjunto_MIME) ## se agrega el archivo al correo

# Configurar conexión SMTP
try:
    sesion_smtp = smtplib.SMTP('smtp.gmail.com', 587) ## Se crea una conexión con el servidor SMTP de Gmail, utilizando el puerto 587 (que admite conexiones seguras con TLS)
    sesion_smtp.starttls()
    sesion_smtp.login(remitente, contra) ## inicio de sesion 
    
    # Convertir el mensaje a formato string y enviar
    texto = mensaje.as_string()
    sesion_smtp.sendmail(remitente, destinatarios, texto)
    print("Correo enviado exitosamente.")
    
except Exception as e:
    print(f"Error enviando el correo: {e}")

finally:
    sesion_smtp.quit() ## se cierra la conexión SMTP
   

